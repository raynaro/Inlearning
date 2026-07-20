import os
import sqlite3
from datetime import datetime, timedelta
from functools import wraps
from zoneinfo import ZoneInfo

try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
    from psycopg2 import IntegrityError as PostgresIntegrityError
except Exception:
    psycopg2 = None
    RealDictCursor = None
    PostgresIntegrityError = Exception

from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, jsonify
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

APP_DIR = os.path.dirname(os.path.abspath(__file__))
DB = os.path.join(APP_DIR, 'database.db')
DATABASE_URL = os.environ.get('DATABASE_URL', '').strip()
USE_POSTGRES = bool(DATABASE_URL)
UPLOADS = os.path.join(APP_DIR, 'uploads')
BASE_XLSX = os.path.join(APP_DIR, 'base.xlsx')
TEACHERS_XLSX = os.path.join(APP_DIR, 'base_docentes.xlsx')
TZ = ZoneInfo('America/Lima')
os.makedirs(UPLOADS, exist_ok=True)

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'cambia-esta-clave-en-produccion')

ROLES = {
    'admin': 'Administrador',
    'sae': 'SAE',
    'commercial': 'Comercial',
    'security': 'Seguridad',
}
AREAS = ('SAE', 'COMERCIAL')
INSTITUTES = ('IDAT', 'ZEGEL')
BRANCH_CAMPUSES = {
    'IDAT': ('SJM', 'SJL', 'ATE', 'PT', 'TV'),
    'ZEGEL': ('SJM', 'SJL', 'ATE'),
}
CAMPUSES = ('SJM', 'SJL', 'ATE', 'PT', 'TV')
BRANCHES = tuple((inst, campus) for inst in INSTITUTES for campus in BRANCH_CAMPUSES[inst])

CAMPUS_CODES = {
    'SJM': 'SJM',
    'SJL': 'SJL',
    'ATE': 'ATE',
    'PT': 'PT',
    'TV': 'TV',
}


class DBCursor:
    def __init__(self, cursor):
        self.cursor = cursor

    def _sql(self, sql):
        return sql.replace('?', '%s') if USE_POSTGRES else sql

    def execute(self, sql, params=()):
        self.cursor.execute(self._sql(sql), params)
        return self

    def fetchone(self):
        return self.cursor.fetchone()

    def fetchall(self):
        return self.cursor.fetchall()


class DBConn:
    def __init__(self):
        if USE_POSTGRES:
            if psycopg2 is None:
                raise RuntimeError('Falta instalar psycopg2-binary para usar PostgreSQL.')
            self.raw = psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)
        else:
            self.raw = sqlite3.connect(DB)
            self.raw.row_factory = sqlite3.Row

    def cursor(self):
        return DBCursor(self.raw.cursor())

    def execute(self, sql, params=()):
        return self.cursor().execute(sql, params)

    def commit(self):
        self.raw.commit()

    def rollback(self):
        self.raw.rollback()

    def close(self):
        self.raw.close()


def conn():
    return DBConn()


DBIntegrityError = (sqlite3.IntegrityError, PostgresIntegrityError)


def now_dt():
    return datetime.now(TZ)


def now_iso():
    return now_dt().isoformat(timespec='seconds')


def today_lima():
    return now_dt().date().isoformat()


def normalize_header(value):
    return str(value or '').strip().lower().replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u')


def only_digits(value):
    return ''.join(ch for ch in str(value or '').replace('.0', '') if ch.isdigit())


def normalize_institute(value, default=''):
    value = str(value or '').strip().upper()
    return value if value in INSTITUTES else default


def normalize_campus(value, default=''):
    value = str(value or '').strip().upper().replace('  ', ' ')
    aliases = {
        'SAN JUAN DE MIRAFLORES': 'SJM',
        'LIMA SUR': 'SJM',
        'SUR': 'SJM',
        'SAN JUAN DE LURIGANCHO': 'SJL',
        'LIMA NORTE': 'SJL',
        'NORTE': 'SJL',
        'LIMA ESTE': 'ATE',
        'ESTE': 'ATE',
        'PURUCHUCO': 'PT',
        'LIMA CENTRO': 'PT',
        'CENTRO': 'PT',
        'CENTRO 1': 'PT',
        'LIMA CENTRO 1': 'PT',
        'TOMAS VALLE': 'TV',
        'TOMÁS VALLE': 'TV',
        'LIMA CENTRO 2': 'TV',
        'CENTRO2': 'TV',
        'CENTRO 2': 'TV',
    }
    value = aliases.get(value, value)
    return value if value in CAMPUSES else default


def valid_branch(institute, campus):
    return (institute, campus) in BRANCHES


def normalize_branch(institute, campus, default_institute='', default_campus=''):
    institute = normalize_institute(institute, default_institute)
    campus = normalize_campus(campus, default_campus)
    if institute and campus and valid_branch(institute, campus):
        return institute, campus
    if default_institute and default_campus and valid_branch(default_institute, default_campus):
        return default_institute, default_campus
    return institute, ''


def branch_label(institute, campus):
    return f'{institute} {campus}' if institute and campus else 'Alcance global'


def session_scope():
    return normalize_institute(session.get('institute')), normalize_campus(session.get('campus'))


def scoped_user():
    return session.get('role') != 'admin'


def scope_filter(alias='', institute='', campus=''):
    prefix = f'{alias}.' if alias else ''
    conditions = []
    params = []
    if institute in INSTITUTES and campus in CAMPUSES and not valid_branch(institute, campus):
        return ['1=0'], []
    if institute in INSTITUTES:
        conditions.append(f'{prefix}institute=?')
        params.append(institute)
    if campus in CAMPUSES:
        conditions.append(f'{prefix}campus=?')
        params.append(campus)
    return conditions, params


def audit(action, target_type='', target_id=None, detail=''):
    if not session.get('user_id'):
        return
    c = conn()
    c.execute('''INSERT INTO audit_logs(admin_user_id,action,target_type,target_id,detail,created_at)
                 VALUES(?,?,?,?,?,?)''',
              (session.get('user_id'), action, target_type, target_id, detail, now_iso()))
    c.commit(); c.close()


def ensure_column(cur, table, column, definition):
    if USE_POSTGRES:
        row = cur.execute('''SELECT column_name FROM information_schema.columns
                             WHERE table_name=? AND column_name=?''', (table, column)).fetchone()
        if not row:
            cur.execute(f'ALTER TABLE {table} ADD COLUMN {column} {definition}')
    else:
        cols = [r['name'] for r in cur.execute(f'PRAGMA table_info({table})').fetchall()]
        if column not in cols:
            cur.execute(f'ALTER TABLE {table} ADD COLUMN {column} {definition}')


def create_index(cur, name, table, columns):
    try:
        cur.execute(f'CREATE INDEX IF NOT EXISTS {name} ON {table}({columns})')
    except Exception:
        pass


def migrate_legacy_scopes(cur):
    """Convierte sedes antiguas a la nueva estructura oficial sin mezclar institutos."""
    idat_map = {
        'LIMA SUR': 'SJM', 'LIMA NORTE': 'SJL', 'LIMA ESTE': 'ATE',
        'LIMA CENTRO': 'PT', 'LIMA CENTRO 2': 'TV',
    }
    zegel_map = {
        'LIMA SUR': 'SJM', 'LIMA NORTE': 'SJL', 'LIMA ESTE': 'ATE',
        'LIMA CENTRO': 'SJM', 'LIMA CENTRO 2': 'SJL', 'PT': 'SJM', 'TV': 'SJL',
    }
    specs = [
        ('users', 'institute', 'campus'), ('students', 'institute', 'campus'),
        ('teachers', 'institute', 'campus'), ('visitors', 'institute', 'campus'),
        ('registration_requests', 'institute', 'campus'),
        ('access_logs', 'registered_institute', 'registered_campus'),
        ('access_logs', 'student_institute', 'student_campus'),
        ('teacher_logs', 'registered_institute', 'registered_campus'),
        ('teacher_logs', 'teacher_institute', 'teacher_campus'),
    ]
    for table, inst_col, campus_col in specs:
        try:
            rows = cur.execute(f'SELECT id,{inst_col} institute_value,{campus_col} campus_value FROM {table}').fetchall()
        except Exception:
            continue
        for row in rows:
            institute = normalize_institute(row['institute_value'], 'IDAT')
            raw_campus = str(row['campus_value'] or '').strip().upper()
            campus = (idat_map if institute == 'IDAT' else zegel_map).get(raw_campus, normalize_campus(raw_campus))
            if not valid_branch(institute, campus):
                campus = BRANCH_CAMPUSES[institute][0]
            cur.execute(f'UPDATE {table} SET {inst_col}=?,{campus_col}=? WHERE id=?', (institute, campus, row['id']))


def assign_operational_scopes(cur):
    """Asigna cada cuenta operativa a una sede oficial y garantiza cobertura mínima."""
    branch_order = list(BRANCHES)

    primary_emails = {
        'security': 'seguridad@idat.edu.pe',
        'sae': 'sae@idat.edu.pe',
        'commercial': 'comercial@idat.edu.pe',
    }

    for role in ('security', 'sae', 'commercial'):
        rows = cur.execute('SELECT id,email,institute,campus FROM users WHERE role=? ORDER BY id', (role,)).fetchall()
        valid_rows = [r for r in rows if valid_branch(r['institute'], r['campus'])]
        unassigned = [r for r in rows if not valid_branch(r['institute'], r['campus'])]
        covered = {(r['institute'], r['campus']) for r in valid_rows}

        primary = next((r for r in unassigned if str(r['email']).lower() == primary_emails[role]), None)
        if primary:
            cur.execute('UPDATE users SET institute=?,campus=? WHERE id=?', ('IDAT', 'SJM', primary['id']))
            covered.add(('IDAT', 'SJM'))
            unassigned = [r for r in unassigned if r['id'] != primary['id']]

        missing = [branch for branch in branch_order if branch not in covered]
        for row, branch in zip(unassigned, missing):
            cur.execute('UPDATE users SET institute=?,campus=? WHERE id=?', (branch[0], branch[1], row['id']))
            covered.add(branch)
        extras = unassigned[len(missing):]
        for index, row in enumerate(extras):
            branch = branch_order[index % len(branch_order)]
            cur.execute('UPDATE users SET institute=?,campus=? WHERE id=?', (branch[0], branch[1], row['id']))

        for institute, campus in branch_order:
            exists = cur.execute('SELECT id FROM users WHERE role=? AND institute=? AND campus=? AND active=1 LIMIT 1',
                                 (role, institute, campus)).fetchone()
            if exists:
                continue
            campus_code = CAMPUS_CODES[campus].lower()
            inst_lower = institute.lower()
            role_prefix = {'security': 'seguridad', 'sae': 'sae', 'commercial': 'comercial'}[role]
            email = f'{role_prefix}.{inst_lower}.{campus_code}@pdr.local'
            code_prefix = {'security': 'SEG', 'sae': 'SAE', 'commercial': 'COM'}[role]
            account_code = f'{code_prefix}-{institute[:3]}-{CAMPUS_CODES[campus]}'
            password = {
                'security': f'Seg{institute.title()}{campus}1',
                'sae': f'Sae{institute.title()}{campus}1',
                'commercial': f'Com{institute.title()}{campus}1',
            }[role]
            name = f'{ROLES[role]} {institute} {campus}'
            cur.execute('INSERT INTO users(name,email,password,role,account_code,institute,campus,active,created_at) VALUES(?,?,?,?,?,?,?,1,?)',
                        (name, email, generate_password_hash(password), role, account_code, institute, campus, now_iso()))


def normalize_operational_account_codes(cur):
    """Genera códigos claros y únicos por instituto, sede y rol para la gestión administrativa."""
    prefixes = {'security': 'SEG', 'sae': 'SAE', 'commercial': 'COM'}
    for institute, campus in BRANCHES:
        for role, prefix in prefixes.items():
            rows = cur.execute('''SELECT id,account_code FROM users
                                  WHERE role=? AND institute=? AND campus=? ORDER BY id''',
                               (role, institute, campus)).fetchall()
            for index, row in enumerate(rows, start=1):
                current = str(row['account_code'] or '')
                if not current or current.startswith(('SEG-', 'SAE-', 'COM-')):
                    code = f'{prefix}-{institute}-{campus}-{index:02d}'
                    cur.execute('UPDATE users SET account_code=? WHERE id=?', (code, row['id']))


def init_db():
    c = conn(); cur = c.cursor()
    id_type = 'SERIAL PRIMARY KEY' if USE_POSTGRES else 'INTEGER PRIMARY KEY AUTOINCREMENT'

    cur.execute(f'''CREATE TABLE IF NOT EXISTS users(
        id {id_type}, name TEXT NOT NULL, email TEXT UNIQUE NOT NULL, password TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'security', account_code TEXT, institute TEXT, campus TEXT,
        active INTEGER NOT NULL DEFAULT 1, created_at TEXT NOT NULL
    )''')
    cur.execute(f'''CREATE TABLE IF NOT EXISTS students(
        id {id_type}, name TEXT NOT NULL, dni TEXT NOT NULL, code TEXT NOT NULL UNIQUE,
        entry_date TEXT, institute TEXT NOT NULL DEFAULT 'IDAT', campus TEXT NOT NULL DEFAULT 'SJM',
        active INTEGER NOT NULL DEFAULT 1, created_at TEXT NOT NULL, updated_at TEXT NOT NULL
    )''')
    cur.execute(f'''CREATE TABLE IF NOT EXISTS teachers(
        id {id_type}, name TEXT NOT NULL, dni TEXT NOT NULL UNIQUE, area TEXT, username TEXT,
        email TEXT, phone TEXT, institute TEXT NOT NULL DEFAULT 'IDAT', campus TEXT NOT NULL DEFAULT 'SJM',
        active INTEGER NOT NULL DEFAULT 1, created_at TEXT NOT NULL, updated_at TEXT NOT NULL
    )''')
    cur.execute(f'''CREATE TABLE IF NOT EXISTS searches(
        id {id_type}, user_id INTEGER, query TEXT NOT NULL, result TEXT NOT NULL,
        student_name TEXT, created_at TEXT NOT NULL
    )''')
    cur.execute(f'''CREATE TABLE IF NOT EXISTS access_logs(
        id {id_type}, user_id INTEGER, student_id INTEGER, query TEXT NOT NULL, result TEXT NOT NULL,
        student_name TEXT, student_dni TEXT, student_code TEXT, student_institute TEXT, student_campus TEXT,
        registered_institute TEXT, registered_campus TEXT, note TEXT, created_at TEXT NOT NULL
    )''')
    cur.execute(f'''CREATE TABLE IF NOT EXISTS teacher_logs(
        id {id_type}, user_id INTEGER, teacher_id INTEGER, query TEXT NOT NULL, result TEXT NOT NULL,
        teacher_name TEXT, teacher_dni TEXT, teacher_area TEXT, teacher_institute TEXT, teacher_campus TEXT,
        registered_institute TEXT, registered_campus TEXT, note TEXT, created_at TEXT NOT NULL
    )''')
    cur.execute(f'''CREATE TABLE IF NOT EXISTS visitors(
        id {id_type}, user_id INTEGER, full_name TEXT NOT NULL, dni TEXT NOT NULL, phone TEXT,
        destination_area TEXT NOT NULL, visit_type TEXT, reason TEXT,
        visit_status TEXT NOT NULL DEFAULT 'NUEVA', attended INTEGER NOT NULL DEFAULT 0,
        attended_at TEXT, attended_by_user_id INTEGER, status_updated_by_user_id INTEGER,
        institute TEXT NOT NULL DEFAULT 'IDAT', campus TEXT NOT NULL DEFAULT 'SJM',
        created_at TEXT NOT NULL
    )''')
    cur.execute(f'''CREATE TABLE IF NOT EXISTS audit_logs(
        id {id_type}, admin_user_id INTEGER, action TEXT NOT NULL, target_type TEXT,
        target_id INTEGER, detail TEXT, created_at TEXT NOT NULL
    )''')
    cur.execute(f'''CREATE TABLE IF NOT EXISTS registration_requests(
        id {id_type}, requester_user_id INTEGER, request_type TEXT NOT NULL DEFAULT 'ALUMNO',
        full_name TEXT NOT NULL, dni TEXT NOT NULL, code TEXT, phone TEXT, detail TEXT,
        institute TEXT NOT NULL DEFAULT 'IDAT', campus TEXT NOT NULL DEFAULT 'SJM',
        status TEXT NOT NULL DEFAULT 'PENDIENTE', reviewed_by_user_id INTEGER,
        reviewed_at TEXT, created_at TEXT NOT NULL
    )''')

    # Migraciones sobre proyectos anteriores.
    ensure_column(cur, 'users', 'role', "TEXT NOT NULL DEFAULT 'security'")
    ensure_column(cur, 'users', 'account_code', 'TEXT')
    ensure_column(cur, 'users', 'institute', 'TEXT')
    ensure_column(cur, 'users', 'campus', 'TEXT')
    ensure_column(cur, 'users', 'active', 'INTEGER NOT NULL DEFAULT 1')
    ensure_column(cur, 'students', 'institute', "TEXT NOT NULL DEFAULT 'IDAT'")
    ensure_column(cur, 'students', 'campus', "TEXT NOT NULL DEFAULT 'SJM'")
    ensure_column(cur, 'teachers', 'username', 'TEXT')
    ensure_column(cur, 'teachers', 'email', 'TEXT')
    ensure_column(cur, 'teachers', 'phone', 'TEXT')
    ensure_column(cur, 'teachers', 'institute', "TEXT NOT NULL DEFAULT 'IDAT'")
    ensure_column(cur, 'teachers', 'campus', "TEXT NOT NULL DEFAULT 'SJM'")
    ensure_column(cur, 'access_logs', 'student_institute', 'TEXT')
    ensure_column(cur, 'access_logs', 'student_campus', 'TEXT')
    ensure_column(cur, 'access_logs', 'registered_institute', 'TEXT')
    ensure_column(cur, 'access_logs', 'registered_campus', 'TEXT')
    ensure_column(cur, 'teacher_logs', 'teacher_institute', 'TEXT')
    ensure_column(cur, 'teacher_logs', 'teacher_campus', 'TEXT')
    ensure_column(cur, 'teacher_logs', 'registered_institute', 'TEXT')
    ensure_column(cur, 'teacher_logs', 'registered_campus', 'TEXT')
    ensure_column(cur, 'visitors', 'phone', 'TEXT')
    ensure_column(cur, 'visitors', 'visit_type', 'TEXT')
    ensure_column(cur, 'visitors', 'visit_status', "TEXT NOT NULL DEFAULT 'NUEVA'")
    ensure_column(cur, 'visitors', 'attended', 'INTEGER NOT NULL DEFAULT 0')
    ensure_column(cur, 'visitors', 'attended_at', 'TEXT')
    ensure_column(cur, 'visitors', 'attended_by_user_id', 'INTEGER')
    ensure_column(cur, 'visitors', 'status_updated_by_user_id', 'INTEGER')
    ensure_column(cur, 'visitors', 'institute', "TEXT NOT NULL DEFAULT 'IDAT'")
    ensure_column(cur, 'visitors', 'campus', "TEXT NOT NULL DEFAULT 'SJM'")

    cur.execute("UPDATE users SET role='commercial' WHERE lower(role) IN ('sales','ventas','comercial')")
    cur.execute("UPDATE visitors SET destination_area='COMERCIAL' WHERE upper(destination_area)='VENTAS'")
    cur.execute("UPDATE visitors SET visit_status='ATENDIDO' WHERE attended=1 AND (visit_status IS NULL OR visit_status='' OR visit_status='NUEVA')")
    cur.execute("UPDATE visitors SET visit_status='NUEVA' WHERE visit_status IS NULL OR visit_status=''")
    cur.execute("UPDATE users SET active=1 WHERE active IS NULL")
    migrate_legacy_scopes(cur)
    cur.execute("UPDATE access_logs SET student_institute=(SELECT institute FROM students WHERE students.id=access_logs.student_id) WHERE (student_institute IS NULL OR student_institute='') AND student_id IS NOT NULL")
    cur.execute("UPDATE access_logs SET student_campus=(SELECT campus FROM students WHERE students.id=access_logs.student_id) WHERE (student_campus IS NULL OR student_campus='') AND student_id IS NOT NULL")

    old_commercial = cur.execute('SELECT id FROM users WHERE lower(email)=?', ('ventas@idat.edu.pe',)).fetchone()
    current_commercial = cur.execute('SELECT id FROM users WHERE lower(email)=?', ('comercial@idat.edu.pe',)).fetchone()
    if old_commercial and not current_commercial:
        cur.execute("UPDATE users SET name='Personal Comercial',email='comercial@idat.edu.pe',role='commercial' WHERE id=?", (old_commercial['id'],))

    main_users = [
        ('Administrador General', 'admin@idat.edu.pe', 'admin123', 'admin', 'ADM-01', None, None),
        ('Personal SAE 01', 'sae@idat.edu.pe', 'sae123', 'sae', 'SAE-01', 'IDAT', 'SJM'),
        ('Personal Comercial', 'comercial@idat.edu.pe', 'comercial123', 'commercial', 'COM-01', 'IDAT', 'SJM'),
        ('Personal de Seguridad 01', 'seguridad@idat.edu.pe', 'seguridad123', 'security', 'SEG-01', 'IDAT', 'SJM'),
    ]
    for name, email, password, role, code, institute, campus in main_users:
        row = cur.execute('SELECT id FROM users WHERE lower(email)=?', (email.lower(),)).fetchone()
        if not row:
            cur.execute('''INSERT INTO users(name,email,password,role,account_code,institute,campus,active,created_at)
                           VALUES(?,?,?,?,?,?,?,1,?)''',
                        (name, email, generate_password_hash(password), role, code, institute, campus, now_iso()))
        else:
            cur.execute('''UPDATE users SET role=?,account_code=COALESCE(account_code,?),
                           institute=COALESCE(institute,?),campus=COALESCE(campus,?) WHERE id=?''',
                        (role, code, institute, campus, row['id']))

    # Conserva las 15 cuentas Seguridad y 20 cuentas SAE de la versión anterior.
    for index in range(2, 16):
        email = f'seguridad{index:02d}@pdr.local'
        exists = cur.execute('SELECT id FROM users WHERE lower(email)=?', (email,)).fetchone()
        if not exists:
            cur.execute('''INSERT INTO users(name,email,password,role,account_code,active,created_at)
                           VALUES(?,?,?,?,?,1,?)''',
                        (f'Personal de Seguridad {index:02d}', email, generate_password_hash(f'Seguridad{index:02d}'), 'security', f'SEG-{index:02d}', now_iso()))
    for index in range(2, 21):
        email = f'sae{index:02d}@pdr.local'
        exists = cur.execute('SELECT id FROM users WHERE lower(email)=?', (email,)).fetchone()
        if not exists:
            cur.execute('''INSERT INTO users(name,email,password,role,account_code,active,created_at)
                           VALUES(?,?,?,?,?,1,?)''',
                        (f'Personal SAE {index:02d}', email, generate_password_hash(f'SAE{index:02d}'), 'sae', f'SAE-{index:02d}', now_iso()))

    assign_operational_scopes(cur)
    normalize_operational_account_codes(cur)

    create_index(cur, 'idx_users_scope', 'users', 'role,institute,campus,active')
    create_index(cur, 'idx_students_scope', 'students', 'institute,campus,active')
    create_index(cur, 'idx_students_dni_code', 'students', 'dni,code')
    create_index(cur, 'idx_visitors_scope', 'visitors', 'institute,campus,destination_area,visit_status')
    create_index(cur, 'idx_visitors_created', 'visitors', 'created_at')
    create_index(cur, 'idx_access_scope', 'access_logs', 'registered_institute,registered_campus,created_at')
    create_index(cur, 'idx_teacher_scope', 'teacher_logs', 'registered_institute,registered_campus,created_at')
    create_index(cur, 'idx_requests_scope', 'registration_requests', 'institute,campus,status,created_at')

    c.commit(); c.close()

    c = conn(); total = c.execute('SELECT COUNT(*) AS n FROM students').fetchone()['n']; c.close()
    if total == 0 and os.path.exists(BASE_XLSX):
        try:
            import_students_from_excel(BASE_XLSX, forced_institute='IDAT', forced_campus='SJM')
        except Exception:
            pass
    c = conn(); total_teachers = c.execute('SELECT COUNT(*) AS n FROM teachers').fetchone()['n']; c.close()
    if total_teachers == 0 and os.path.exists(TEACHERS_XLSX):
        try:
            import_teachers_from_excel(TEACHERS_XLSX, forced_institute='IDAT', forced_campus='SJM')
        except Exception:
            pass
    purge_old_logs()


def purge_old_logs():
    # Retención ampliada para reportes históricos y operación a mayor escala.
    limit = (now_dt() - timedelta(days=365)).isoformat(timespec='seconds')
    c = conn()
    c.execute('DELETE FROM searches WHERE created_at < ?', (limit,))
    c.execute('DELETE FROM access_logs WHERE created_at < ?', (limit,))
    c.execute('DELETE FROM teacher_logs WHERE created_at < ?', (limit,))
    c.execute('DELETE FROM visitors WHERE created_at < ?', (limit,))
    c.commit(); c.close()


def import_students_from_excel(path, forced_institute=None, forced_campus=None):
    forced_institute = normalize_institute(forced_institute)
    forced_campus = normalize_campus(forced_campus)
    if forced_institute and forced_campus and not valid_branch(forced_institute, forced_campus):
        raise ValueError('La combinación de instituto y sede no es válida.')
    wb = load_workbook(path, data_only=True)
    records = []
    targets = set()

    def find_index(headers, *names):
        names = [normalize_header(n) for n in names]
        return next((headers.index(n) for n in names if n in headers), None)

    for ws in wb.worksheets:
        headers = [normalize_header(c.value) for c in ws[1]]
        i_name = find_index(headers, 'NombreCompleto', 'Nombre Completo', 'Nombre', 'Alumno')
        i_dni = find_index(headers, 'DNI', 'Documento')
        i_code = find_index(headers, 'Codigo', 'Código', 'Code')
        i_date = find_index(headers, 'Fecha Ingreso', 'Fecha de Ingreso', 'Fecha')
        i_inst = find_index(headers, 'INSTITUTO', 'Instituto', 'Institución', 'Institucion')
        i_campus = find_index(headers, 'SEDE', 'Sede', 'Campus', 'Local')
        if i_name is None or i_dni is None or i_code is None:
            continue
        title = str(ws.title or '').upper()
        sheet_inst = next((i for i in INSTITUTES if i in title), '')
        sheet_campus = next((s for s in CAMPUSES if s in title), '')
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = str(row[i_name] or '').strip().upper()
            dni = only_digits(row[i_dni])
            code = str(row[i_code] or '').strip().replace('.0', '').upper()
            entry_date = str(row[i_date] or '').strip() if i_date is not None else now_dt().strftime('%d/%m/%Y')
            row_inst = normalize_institute(row[i_inst]) if i_inst is not None else ''
            row_campus = normalize_campus(row[i_campus]) if i_campus is not None else ''
            institute = forced_institute or row_inst or sheet_inst or 'IDAT'
            campus = forced_campus or row_campus or sheet_campus or 'SJM'
            if not name or not dni or not code or not valid_branch(institute, campus):
                continue
            records.append((name, dni, code, entry_date, institute, campus))
            targets.add((institute, campus))

    if not records:
        raise ValueError('No se encontraron alumnos válidos. El Excel debe incluir Nombre, DNI y Código.')

    c = conn(); cur = c.cursor(); stamp = now_iso()
    for institute, campus in targets:
        cur.execute('UPDATE students SET active=0,updated_at=? WHERE institute=? AND campus=?', (stamp, institute, campus))
    counts = {}
    for name, dni, code, entry_date, institute, campus in records:
        cur.execute('''INSERT INTO students(name,dni,code,entry_date,institute,campus,active,created_at,updated_at)
                       VALUES(?,?,?,?,?,?,1,?,?)
                       ON CONFLICT(code) DO UPDATE SET name=excluded.name,dni=excluded.dni,
                       entry_date=excluded.entry_date,institute=excluded.institute,campus=excluded.campus,
                       active=1,updated_at=excluded.updated_at''',
                    (name, dni, code, entry_date, institute, campus, stamp, stamp))
        counts[(institute, campus)] = counts.get((institute, campus), 0) + 1
    c.commit(); c.close()
    return counts


def import_teachers_from_excel(path, forced_institute=None, forced_campus=None):
    institute = normalize_institute(forced_institute, 'IDAT')
    campus = normalize_campus(forced_campus, 'SJM')
    if not valid_branch(institute, campus):
        raise ValueError('La combinación de instituto y sede no es válida.')
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [normalize_header(c.value) for c in ws[1]]

    def find(*names):
        names = [normalize_header(n) for n in names]
        return next((headers.index(n) for n in names if n in headers), None)

    i_name = find('NombreCompleto', 'Nombre Completo', 'Nombre', 'Docente')
    i_dni = find('DNI', 'Documento')
    i_area = find('Area', 'Área', 'Especialidad', 'Curso', 'Marca')
    i_user = find('Usuario', 'User')
    i_email = find('Correo', 'Email', 'E-mail')
    i_phone = find('Telefono Docente', 'Telefono', 'Teléfono', 'Celular')
    i_inst = find('Instituto', 'Institución')
    i_campus = find('Sede', 'Campus', 'Local')
    if i_name is None or i_dni is None:
        raise ValueError('El Excel de docentes debe incluir Nombre y DNI.')

    c = conn(); cur = c.cursor(); stamp = now_iso(); count = 0
    cur.execute('UPDATE teachers SET active=0,updated_at=? WHERE institute=? AND campus=?', (stamp, institute, campus))
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[i_name] or '').strip().upper()
        dni = only_digits(row[i_dni])
        if not name or not dni:
            continue
        row_inst = normalize_institute(row[i_inst], institute) if i_inst is not None else institute
        row_campus = normalize_campus(row[i_campus], campus) if i_campus is not None else campus
        if not valid_branch(row_inst, row_campus):
            continue
        area = str(row[i_area] or '').strip().upper() if i_area is not None else ''
        username = str(row[i_user] or '').strip() if i_user is not None else ''
        email = str(row[i_email] or '').strip() if i_email is not None else ''
        phone = only_digits(row[i_phone]) if i_phone is not None else ''
        cur.execute('''INSERT INTO teachers(name,dni,area,username,email,phone,institute,campus,active,created_at,updated_at)
                       VALUES(?,?,?,?,?,?,?,?,1,?,?)
                       ON CONFLICT(dni) DO UPDATE SET name=excluded.name,area=excluded.area,
                       username=excluded.username,email=excluded.email,phone=excluded.phone,
                       institute=excluded.institute,campus=excluded.campus,active=1,updated_at=excluded.updated_at''',
                    (name, dni, area, username, email, phone, row_inst, row_campus, stamp, stamp))
        count += 1
    c.commit(); c.close()
    return count


def build_activity_charts(c, area, institute='', campus=''):
    charts = {}
    short_days = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']
    end_date = now_dt().date()
    for days in (7, 30):
        start_date = end_date - timedelta(days=days - 1)
        conditions = ['destination_area=?', 'substr(created_at,1,10) BETWEEN ? AND ?']
        params = [area, start_date.isoformat(), end_date.isoformat()]
        extra, extra_params = scope_filter('', institute, campus)
        conditions.extend(extra); params.extend(extra_params)
        rows = c.execute(f'''SELECT substr(created_at,1,10) fecha,COUNT(*) visitas,
                            SUM(CASE WHEN visit_status='ATENDIDO' OR attended=1 THEN 1 ELSE 0 END) atendidos
                            FROM visitors WHERE {' AND '.join(conditions)}
                            GROUP BY substr(created_at,1,10) ORDER BY fecha''', tuple(params)).fetchall()
        by_date = {r['fecha']: r for r in rows}
        labels, full_labels, visits_values, attended_values = [], [], [], []
        for index in range(days):
            current = start_date + timedelta(days=index)
            row = by_date.get(current.isoformat())
            labels.append(f'{short_days[current.weekday()]} {current.day:02d}')
            full_labels.append(current.strftime('%d/%m/%Y'))
            visits_values.append(int(row['visitas'] or 0) if row else 0)
            attended_values.append(int(row['atendidos'] or 0) if row else 0)
        charts[str(days)] = {
            'labels': labels, 'full_labels': full_labels,
            'visits': visits_values, 'attended': attended_values,
            'total_visits': sum(visits_values), 'total_attended': sum(attended_values),
        }
    return charts


def login_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return wrapped


def role_home_endpoint(role):
    return {'admin': 'dashboard', 'security': 'security_panel', 'sae': 'visitors', 'commercial': 'visitors'}.get(role, 'login')


def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            if session.get('role') not in roles:
                flash('No tienes permiso para entrar a esa sección.', 'danger')
                return redirect(url_for(role_home_endpoint(session.get('role'))))
            return f(*args, **kwargs)
        return wrapped
    return decorator


def admin_required(f):
    return role_required('admin')(f)


@app.context_processor
def inject_globals():
    return dict(roles=ROLES, areas=AREAS, institutes=INSTITUTES, campuses=CAMPUSES, branches=BRANCHES, branch_campuses=BRANCH_CAMPUSES, branch_label=branch_label)


@app.route('/')
def home():
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        c = conn(); user = c.execute('SELECT * FROM users WHERE lower(email)=? AND active=1', (email,)).fetchone(); c.close()
        if user and check_password_hash(user['password'], password):
            if user['role'] != 'admin' and (not valid_branch(user['institute'], user['campus'])):
                flash('La cuenta todavía no tiene instituto y sede asignados. Comunícate con Administración.', 'danger')
                return render_template('login.html')
            session.clear()
            session.update(user_id=user['id'], name=user['name'], role=user['role'],
                           institute=user['institute'], campus=user['campus'], account_code=user['account_code'])
            return redirect(url_for(role_home_endpoint(user['role'])))
        flash('Credenciales incorrectas.', 'danger')
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/register', methods=['GET', 'POST'])
@login_required
@admin_required
def register():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        role = request.form.get('role', 'security')
        institute = normalize_institute(request.form.get('institute'))
        campus = normalize_campus(request.form.get('campus'))
        if role not in ROLES:
            role = 'security'
        if role == 'admin':
            institute = campus = None
        if not name or not email or not password:
            flash('Completa nombre, correo y contraseña.', 'danger')
            return render_template('register.html')
        if role != 'admin' and not valid_branch(institute, campus):
            flash('Selecciona una combinación válida de instituto y sede para la cuenta operativa.', 'danger')
            return render_template('register.html')
        try:
            c = conn()
            c.execute('''INSERT INTO users(name,email,password,role,account_code,institute,campus,active,created_at)
                         VALUES(?,?,?,?,?,?,?,1,?)''',
                      (name, email, generate_password_hash(password), role, request.form.get('account_code', '').strip() or None,
                       institute, campus, now_iso()))
            c.commit(); user = c.execute('SELECT id FROM users WHERE lower(email)=?', (email,)).fetchone(); c.close()
            audit('CREAR_USUARIO', 'user', user['id'] if user else None, f'{name} · {role} · {branch_label(institute, campus)}')
            flash('Usuario creado correctamente.', 'success')
            return redirect(url_for('users', role=role, institute=institute or '', campus=campus or ''))
        except DBIntegrityError:
            flash('Ese correo ya existe.', 'danger')
    return render_template('register.html')


@app.route('/usuarios')
@login_required
@admin_required
def users():
    q = request.args.get('q', '').strip()
    role_filter = request.args.get('role', '').strip()
    institute_filter = normalize_institute(request.args.get('institute'))
    campus_filter = normalize_campus(request.args.get('campus'))
    conditions, params = [], []
    if q:
        conditions.append('(name LIKE ? OR email LIKE ? OR account_code LIKE ?)')
        params.extend([f'%{q}%', f'%{q}%', f'%{q}%'])
    if role_filter in ROLES:
        conditions.append('role=?'); params.append(role_filter)
    if institute_filter:
        conditions.append('institute=?'); params.append(institute_filter)
    if campus_filter:
        conditions.append('campus=?'); params.append(campus_filter)
    where = ' WHERE ' + ' AND '.join(conditions) if conditions else ''
    c = conn()
    data = c.execute('''SELECT id,name,email,role,account_code,institute,campus,active,created_at
                        FROM users''' + where + ' ORDER BY institute,campus,role,account_code,name', tuple(params)).fetchall()
    counts = {role: c.execute('SELECT COUNT(*) n FROM users WHERE role=?', (role,)).fetchone()['n'] for role in ROLES}
    counts['inactive'] = c.execute('SELECT COUNT(*) n FROM users WHERE active=0').fetchone()['n']
    branch_counts = c.execute('''SELECT institute,campus,role,COUNT(*) total
                                 FROM users WHERE role<>'admin' GROUP BY institute,campus,role''').fetchall()
    c.close()
    return render_template('users.html', users=data, q=q, role_filter=role_filter,
                           institute_filter=institute_filter, campus_filter=campus_filter,
                           counts=counts, branch_counts=branch_counts)


@app.route('/usuarios/<int:user_id>/editar', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_user(user_id):
    c = conn(); user = c.execute('SELECT * FROM users WHERE id=?', (user_id,)).fetchone()
    if not user:
        c.close(); flash('Usuario no encontrado.', 'danger'); return redirect(url_for('users'))
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip().lower()
        role = request.form.get('role', '').strip()
        password = request.form.get('password', '')
        active = 1 if request.form.get('active') == '1' else 0
        institute = normalize_institute(request.form.get('institute'))
        campus = normalize_campus(request.form.get('campus'))
        account_code = request.form.get('account_code', '').strip() or None
        if role == 'admin':
            institute = campus = None
        if not name or not email or role not in ROLES:
            c.close(); flash('Completa nombre, correo y rol válidos.', 'danger'); return render_template('edit_user.html', user=user)
        if role != 'admin' and not valid_branch(institute, campus):
            c.close(); flash('Selecciona una combinación válida de instituto y sede.', 'danger'); return render_template('edit_user.html', user=user)
        if user_id == session.get('user_id') and active == 0:
            c.close(); flash('No puedes desactivar tu propia cuenta.', 'danger'); return render_template('edit_user.html', user=user)
        try:
            if password:
                c.execute('''UPDATE users SET name=?,email=?,role=?,account_code=?,institute=?,campus=?,active=?,password=? WHERE id=?''',
                          (name, email, role, account_code, institute, campus, active, generate_password_hash(password), user_id))
            else:
                c.execute('''UPDATE users SET name=?,email=?,role=?,account_code=?,institute=?,campus=?,active=? WHERE id=?''',
                          (name, email, role, account_code, institute, campus, active, user_id))
            c.commit(); c.close()
            audit('EDITAR_USUARIO', 'user', user_id, f'{name} · {role} · {branch_label(institute, campus)} · activo={active}')
            flash('Cuenta actualizada correctamente.', 'success')
            return redirect(url_for('users', role=role, institute=institute or '', campus=campus or ''))
        except DBIntegrityError:
            c.close(); flash('Ese correo ya está utilizado por otra cuenta.', 'danger')
    else:
        c.close()
    return render_template('edit_user.html', user=user)


@app.route('/sedes')
@login_required
@admin_required
def branches():
    c = conn(); today = today_lima(); cards = []
    for institute, campus in BRANCHES:
        card = {
            'institute': institute, 'campus': campus,
            'students': c.execute('SELECT COUNT(*) n FROM students WHERE active=1 AND institute=? AND campus=?', (institute, campus)).fetchone()['n'],
            'security': c.execute("SELECT COUNT(*) n FROM users WHERE active=1 AND role='security' AND institute=? AND campus=?", (institute, campus)).fetchone()['n'],
            'sae': c.execute("SELECT COUNT(*) n FROM users WHERE active=1 AND role='sae' AND institute=? AND campus=?", (institute, campus)).fetchone()['n'],
            'commercial': c.execute("SELECT COUNT(*) n FROM users WHERE active=1 AND role='commercial' AND institute=? AND campus=?", (institute, campus)).fetchone()['n'],
            'visits_today': c.execute("SELECT COUNT(*) n FROM visitors WHERE institute=? AND campus=? AND substr(created_at,1,10)=?", (institute, campus, today)).fetchone()['n'],
            'pending': c.execute("SELECT COUNT(*) n FROM visitors WHERE institute=? AND campus=? AND visit_status='NUEVA'", (institute, campus)).fetchone()['n'],
            'registration_requests': c.execute("SELECT COUNT(*) n FROM registration_requests WHERE institute=? AND campus=? AND status='PENDIENTE'", (institute, campus)).fetchone()['n'],
        }
        cards.append(card)
    audits = c.execute('''SELECT a.*,u.name admin_name FROM audit_logs a LEFT JOIN users u ON u.id=a.admin_user_id
                          ORDER BY a.id DESC LIMIT 12''').fetchall()
    c.close()
    return render_template('branches.html', branch_cards=cards, audits=audits)


@app.route('/sedes/<institute>/<path:campus>')
@login_required
@admin_required
def branch_detail(institute, campus):
    institute = normalize_institute(institute)
    campus = normalize_campus(campus)
    if not valid_branch(institute, campus):
        flash('Instituto o sede inválidos.', 'danger'); return redirect(url_for('branches'))
    c = conn(); today = today_lima()
    stats = {
        'students': c.execute('SELECT COUNT(*) n FROM students WHERE active=1 AND institute=? AND campus=?', (institute, campus)).fetchone()['n'],
        'users': c.execute("SELECT COUNT(*) n FROM users WHERE active=1 AND role<>'admin' AND institute=? AND campus=?", (institute, campus)).fetchone()['n'],
        'visits_today': c.execute("SELECT COUNT(*) n FROM visitors WHERE institute=? AND campus=? AND substr(created_at,1,10)=?", (institute, campus, today)).fetchone()['n'],
        'pending': c.execute("SELECT COUNT(*) n FROM visitors WHERE institute=? AND campus=? AND visit_status='NUEVA'", (institute, campus)).fetchone()['n'],
        'attended': c.execute("SELECT COUNT(*) n FROM visitors WHERE institute=? AND campus=? AND visit_status='ATENDIDO'", (institute, campus)).fetchone()['n'],
        'registration_requests': c.execute("SELECT COUNT(*) n FROM registration_requests WHERE institute=? AND campus=? AND status='PENDIENTE'", (institute, campus)).fetchone()['n'],
    }
    users_data = c.execute('''SELECT * FROM users WHERE role<>'admin' AND institute=? AND campus=? ORDER BY role,name''', (institute, campus)).fetchall()
    visits_data = c.execute('''SELECT v.*,ru.name registered_by,au.name attended_by FROM visitors v
                               LEFT JOIN users ru ON ru.id=v.user_id LEFT JOIN users au ON au.id=v.attended_by_user_id
                               WHERE v.institute=? AND v.campus=? ORDER BY v.id DESC LIMIT 20''', (institute, campus)).fetchall()
    charts = {'SAE': build_activity_charts(c, 'SAE', institute, campus), 'COMERCIAL': build_activity_charts(c, 'COMERCIAL', institute, campus)}
    requests_data = c.execute('''SELECT r.*,u.name requester_name FROM registration_requests r
                                 LEFT JOIN users u ON u.id=r.requester_user_id
                                 WHERE r.institute=? AND r.campus=? ORDER BY r.id DESC LIMIT 10''',
                              (institute, campus)).fetchall()
    c.close()
    return render_template('branch_detail.html', institute=institute, campus=campus, stats=stats,
                           branch_users=users_data, branch_visits=visits_data, branch_charts=charts,
                           branch_requests=requests_data)


@app.route('/solicitudes')
@login_required
@role_required('admin', 'security')
def registration_requests():
    status_filter = request.args.get('status', '').strip().upper()
    if status_filter not in ('PENDIENTE', 'APROBADA', 'RECHAZADA'):
        status_filter = ''
    if session.get('role') == 'admin':
        institute = normalize_institute(request.args.get('institute'))
        campus = normalize_campus(request.args.get('campus'))
        if institute and campus and not valid_branch(institute, campus):
            campus = ''
    else:
        institute, campus = session_scope()

    conditions, params = [], []
    if institute:
        conditions.append('r.institute=?'); params.append(institute)
    if campus:
        conditions.append('r.campus=?'); params.append(campus)
    if status_filter:
        conditions.append('r.status=?'); params.append(status_filter)
    where = ' WHERE ' + ' AND '.join(conditions) if conditions else ''
    c = conn()
    data = c.execute('''SELECT r.*,u.name requester_name,rv.name reviewer_name
                        FROM registration_requests r
                        LEFT JOIN users u ON u.id=r.requester_user_id
                        LEFT JOIN users rv ON rv.id=r.reviewed_by_user_id''' + where +
                     " ORDER BY CASE r.status WHEN 'PENDIENTE' THEN 0 WHEN 'APROBADA' THEN 1 ELSE 2 END,r.id DESC",
                     tuple(params)).fetchall()
    base_conditions, base_params = [], []
    if institute:
        base_conditions.append('institute=?'); base_params.append(institute)
    if campus:
        base_conditions.append('campus=?'); base_params.append(campus)
    base_where = ' WHERE ' + ' AND '.join(base_conditions) if base_conditions else ''
    counts = {}
    for state in ('PENDIENTE', 'APROBADA', 'RECHAZADA'):
        state_where = base_where + (' AND ' if base_where else ' WHERE ') + 'status=?'
        counts[state.lower()] = c.execute('SELECT COUNT(*) n FROM registration_requests' + state_where,
                                          tuple(base_params + [state])).fetchone()['n']
    c.close()
    return render_template('registration_requests.html', requests_data=data, counts=counts,
                           selected_institute=institute, selected_campus=campus,
                           status_filter=status_filter)


@app.route('/solicitudes/nueva', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'security')
def new_registration_request():
    if session.get('role') == 'admin':
        institute = normalize_institute(request.form.get('institute') if request.method == 'POST' else request.args.get('institute'), 'IDAT')
        campus = normalize_campus(request.form.get('campus') if request.method == 'POST' else request.args.get('campus'), 'SJM')
        if not valid_branch(institute, campus):
            institute, campus = 'IDAT', 'SJM'
    else:
        institute, campus = session_scope()

    if request.method == 'POST':
        request_type = request.form.get('request_type', 'ALUMNO').strip().upper()
        full_name = request.form.get('full_name', '').strip().upper()
        dni = only_digits(request.form.get('dni', ''))
        code = str(request.form.get('code', '') or '').strip().upper()
        phone = only_digits(request.form.get('phone', ''))
        detail = request.form.get('detail', '').strip()
        if request_type not in ('ALUMNO', 'DOCENTE'):
            request_type = 'ALUMNO'
        if not full_name or not dni:
            flash('Completa nombre y DNI.', 'danger')
        elif request_type == 'ALUMNO' and not code:
            flash('Para alumnos debes indicar el código.', 'danger')
        else:
            c = conn()
            c.execute('''INSERT INTO registration_requests(requester_user_id,request_type,full_name,dni,code,phone,
                         detail,institute,campus,status,created_at) VALUES(?,?,?,?,?,?,?,?,?,'PENDIENTE',?)''',
                      (session['user_id'], request_type, full_name, dni, code or None, phone or None,
                       detail or None, institute, campus, now_iso()))
            c.commit(); c.close()
            flash(f'Solicitud enviada a la bandeja de {institute} {campus}.', 'success')
            return redirect(url_for('registration_requests', institute=institute, campus=campus))
    return render_template('registration_request_form.html', institute=institute, campus=campus,
                           prefill_dni=request.args.get('dni', ''))


@app.post('/solicitudes/<int:request_id>/resolver')
@login_required
@admin_required
def resolve_registration_request(request_id):
    decision = request.form.get('decision', '').strip().upper()
    if decision not in ('APROBADA', 'RECHAZADA'):
        flash('Decisión inválida.', 'danger'); return redirect(url_for('registration_requests'))
    c = conn(); item = c.execute('SELECT * FROM registration_requests WHERE id=?', (request_id,)).fetchone()
    if not item:
        c.close(); flash('La solicitud no existe.', 'danger'); return redirect(url_for('registration_requests'))
    if decision == 'APROBADA':
        stamp = now_iso()
        if item['request_type'] == 'ALUMNO':
            if not item['code']:
                c.close(); flash('La solicitud no tiene código de alumno.', 'danger'); return redirect(url_for('registration_requests'))
            c.execute('''INSERT INTO students(name,dni,code,entry_date,institute,campus,active,created_at,updated_at)
                         VALUES(?,?,?,?,?,?,1,?,?)
                         ON CONFLICT(code) DO UPDATE SET name=excluded.name,dni=excluded.dni,
                         institute=excluded.institute,campus=excluded.campus,active=1,updated_at=excluded.updated_at''',
                      (item['full_name'], item['dni'], item['code'], now_dt().strftime('%d/%m/%Y'),
                       item['institute'], item['campus'], stamp, stamp))
        else:
            c.execute('''INSERT INTO teachers(name,dni,area,username,email,phone,institute,campus,active,created_at,updated_at)
                         VALUES(?,?,?,?,?,?,?,?,1,?,?)
                         ON CONFLICT(dni) DO UPDATE SET name=excluded.name,phone=excluded.phone,
                         institute=excluded.institute,campus=excluded.campus,active=1,updated_at=excluded.updated_at''',
                      (item['full_name'], item['dni'], '', '', '', item['phone'] or '',
                       item['institute'], item['campus'], stamp, stamp))
    c.execute('''UPDATE registration_requests SET status=?,reviewed_by_user_id=?,reviewed_at=? WHERE id=?''',
              (decision, session['user_id'], now_iso(), request_id))
    c.commit(); c.close()
    audit('RESOLVER_SOLICITUD', 'registration_request', request_id,
          f'{decision} · {item["request_type"]} · {branch_label(item["institute"], item["campus"])}')
    flash('Solicitud actualizada correctamente.', 'success')
    return redirect(url_for('registration_requests', institute=item['institute'], campus=item['campus']))


@app.route('/seguridad')
@login_required
@role_required('admin', 'security')
def security_panel():
    today = today_lima()
    institute, campus = session_scope()
    if session.get('role') == 'admin':
        institute = normalize_institute(request.args.get('institute'), 'IDAT')
        campus = normalize_campus(request.args.get('campus'), 'SJM')
        if not valid_branch(institute, campus):
            institute, campus = 'IDAT', 'SJM'
    c = conn()
    stats = {
        'students_today': c.execute('''SELECT COUNT(*) n FROM access_logs WHERE registered_institute=? AND registered_campus=? AND substr(created_at,1,10)=?''', (institute, campus, today)).fetchone()['n'],
        'teachers_today': c.execute('''SELECT COUNT(*) n FROM teacher_logs WHERE registered_institute=? AND registered_campus=? AND substr(created_at,1,10)=?''', (institute, campus, today)).fetchone()['n'],
        'visits_today': c.execute('''SELECT COUNT(*) n FROM visitors WHERE institute=? AND campus=? AND substr(created_at,1,10)=?''', (institute, campus, today)).fetchone()['n'],
        'requests_pending': c.execute("SELECT COUNT(*) n FROM registration_requests WHERE institute=? AND campus=? AND status='PENDIENTE'", (institute, campus)).fetchone()['n'],
        'my_actions': c.execute('''SELECT (SELECT COUNT(*) FROM access_logs WHERE user_id=? AND substr(created_at,1,10)=?) +
                                         (SELECT COUNT(*) FROM teacher_logs WHERE user_id=? AND substr(created_at,1,10)=?) +
                                         (SELECT COUNT(*) FROM visitors WHERE user_id=? AND substr(created_at,1,10)=?) +
                                         (SELECT COUNT(*) FROM registration_requests WHERE requester_user_id=? AND substr(created_at,1,10)=?) n''',
                                      (session['user_id'], today, session['user_id'], today, session['user_id'], today,
                                       session['user_id'], today)).fetchone()['n'],
    }
    recent_students = c.execute('''SELECT * FROM access_logs WHERE registered_institute=? AND registered_campus=? ORDER BY id DESC LIMIT 6''', (institute, campus)).fetchall()
    recent_teachers = c.execute('''SELECT * FROM teacher_logs WHERE registered_institute=? AND registered_campus=? ORDER BY id DESC LIMIT 6''', (institute, campus)).fetchall()
    recent_visits = c.execute('''SELECT * FROM visitors WHERE institute=? AND campus=? ORDER BY id DESC LIMIT 6''', (institute, campus)).fetchall()
    c.close()
    return render_template('security_panel.html', stats=stats, recent_students=recent_students,
                           recent_teachers=recent_teachers, recent_visits=recent_visits,
                           institute=institute, campus=campus)


@app.route('/consulta', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'security')
def consulta():
    result = None; q = ''; status = None
    if session.get('role') == 'admin':
        institute = normalize_institute(request.form.get('institute') if request.method == 'POST' else request.args.get('institute'), 'IDAT')
        campus = normalize_campus(request.form.get('campus') if request.method == 'POST' else request.args.get('campus'), 'SJM')
        if not valid_branch(institute, campus):
            institute, campus = 'IDAT', 'SJM'
    else:
        institute, campus = session_scope()
    if request.method == 'POST':
        q = request.form.get('q', '').strip().upper()
        note = request.form.get('note', '').strip()
        c = conn()
        student = c.execute('''SELECT * FROM students WHERE (dni=? OR code=?) AND institute=? AND campus=?
                               ORDER BY active DESC LIMIT 1''', (q, q, institute, campus)).fetchone()
        status = 'ACTIVO' if student and student['active'] else 'INACTIVO'
        stamp = now_iso()
        c.execute('INSERT INTO searches(user_id,query,result,student_name,created_at) VALUES(?,?,?,?,?)',
                  (session['user_id'], q, status, student['name'] if student else None, stamp))
        c.execute('''INSERT INTO access_logs(user_id,student_id,query,result,student_name,student_dni,student_code,
                     student_institute,student_campus,registered_institute,registered_campus,note,created_at)
                     VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                  (session['user_id'], student['id'] if student else None, q, status,
                   student['name'] if student else None, student['dni'] if student else None,
                   student['code'] if student else None, student['institute'] if student else institute,
                   student['campus'] if student else campus, institute, campus, note, stamp))
        c.commit(); c.close(); result = student
    return render_template('consulta.html', result=result, q=q, status=status, institute=institute, campus=campus)


@app.route('/docentes/registro', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'security')
def teacher_check():
    result = None; q = ''; status = None
    if session.get('role') == 'admin':
        institute = normalize_institute(request.form.get('institute') if request.method == 'POST' else request.args.get('institute'), 'IDAT')
        campus = normalize_campus(request.form.get('campus') if request.method == 'POST' else request.args.get('campus'), 'SJM')
        if not valid_branch(institute, campus):
            institute, campus = 'IDAT', 'SJM'
    else:
        institute, campus = session_scope()
    if request.method == 'POST':
        q = request.form.get('q', '').strip().upper()
        note = request.form.get('note', '').strip()
        c = conn()
        teacher = c.execute('''SELECT * FROM teachers WHERE dni=? AND institute=? AND campus=?''', (q, institute, campus)).fetchone()
        status = 'ACTIVO' if teacher and teacher['active'] else 'INACTIVO'
        stamp = now_iso()
        c.execute('''INSERT INTO teacher_logs(user_id,teacher_id,query,result,teacher_name,teacher_dni,teacher_area,
                     teacher_institute,teacher_campus,registered_institute,registered_campus,note,created_at)
                     VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                  (session['user_id'], teacher['id'] if teacher else None, q, status,
                   teacher['name'] if teacher else None, teacher['dni'] if teacher else None,
                   teacher['area'] if teacher else None, teacher['institute'] if teacher else institute,
                   teacher['campus'] if teacher else campus, institute, campus, note, stamp))
        c.commit(); c.close(); result = teacher
    return render_template('teacher_check.html', result=result, q=q, status=status, institute=institute, campus=campus)


@app.route('/visitas/nueva', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'security')
def new_visitor():
    if session.get('role') == 'admin':
        institute = normalize_institute(request.form.get('institute') if request.method == 'POST' else request.args.get('institute'), 'IDAT')
        campus = normalize_campus(request.form.get('campus') if request.method == 'POST' else request.args.get('campus'), 'SJM')
        if not valid_branch(institute, campus):
            institute, campus = 'IDAT', 'SJM'
    else:
        institute, campus = session_scope()
    selected_area = request.form.get('destination_area', '').strip().upper() if request.method == 'POST' else request.args.get('area', '').strip().upper()
    if selected_area not in AREAS:
        selected_area = ''
    if request.method == 'POST':
        full_name = request.form.get('full_name', '').strip().upper()
        dni = only_digits(request.form.get('dni', ''))
        phone = only_digits(request.form.get('phone', ''))
        destination_area = selected_area
        reason = request.form.get('reason', '').strip()
        commercial_reason = request.form.get('commercial_reason', '').strip().upper()
        visit_type = ''
        if destination_area not in AREAS:
            flash('Selecciona el área SAE o Comercial.', 'danger')
            return render_template('visitor_form.html', selected_area=destination_area, institute=institute, campus=campus)
        if not full_name or not dni or not phone:
            flash('Completa nombres, DNI y teléfono.', 'danger')
            return render_template('visitor_form.html', selected_area=destination_area, institute=institute, campus=campus)
        if destination_area == 'COMERCIAL':
            if commercial_reason not in ('CARRERA', 'CURSO'):
                flash('Selecciona Carrera o Curso.', 'danger')
                return render_template('visitor_form.html', selected_area=destination_area, institute=institute, campus=campus)
            reason = commercial_reason; visit_type = commercial_reason
        c = conn()
        c.execute('''INSERT INTO visitors(user_id,full_name,dni,phone,destination_area,visit_type,reason,
                     visit_status,institute,campus,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?)''',
                  (session['user_id'], full_name, dni, phone, destination_area, visit_type, reason,
                   'NUEVA', institute, campus, now_iso()))
        c.commit(); c.close()
        flash(f'Visita registrada en {branch_label(institute, campus)}.', 'success')
        return redirect(url_for('security_panel'))
    return render_template('visitor_form.html', selected_area=selected_area, institute=institute, campus=campus)


@app.route('/sae')
@login_required
@role_required('admin', 'sae')
def sae_area():
    return redirect(url_for('visitors', area='SAE'))


@app.route('/comercial')
@login_required
@role_required('admin', 'commercial')
def commercial_area():
    return redirect(url_for('visitors', area='COMERCIAL'))


@app.route('/visitas')
@login_required
@role_required('admin', 'sae', 'commercial')
def visitors():
    role = session.get('role')
    status_filter = request.args.get('status', '').strip().upper()
    if status_filter not in ('NUEVA', 'ATENDIDO', 'NO ATENDIDO'):
        status_filter = ''
    conditions, params = [], []
    area_actual = ''
    if role == 'sae':
        area_actual = 'SAE'; institute, campus = session_scope()
        conditions.extend(['v.destination_area=?', 'v.institute=?', 'v.campus=?'])
        params.extend([area_actual, institute, campus])
    elif role == 'commercial':
        area_actual = 'COMERCIAL'; institute, campus = session_scope()
        conditions.extend(['v.destination_area=?', 'v.institute=?', 'v.campus=?'])
        params.extend([area_actual, institute, campus])
    else:
        area_actual = request.args.get('area', '').strip().upper()
        if area_actual in AREAS:
            conditions.append('v.destination_area=?'); params.append(area_actual)
        else:
            area_actual = ''
        institute = normalize_institute(request.args.get('institute'))
        campus = normalize_campus(request.args.get('campus'))
        extra, extra_params = scope_filter('v', institute, campus)
        conditions.extend(extra); params.extend(extra_params)
    if status_filter:
        conditions.append('v.visit_status=?'); params.append(status_filter)
    where = ' WHERE ' + ' AND '.join(conditions) if conditions else ''
    order = " ORDER BY CASE v.visit_status WHEN 'NUEVA' THEN 1 WHEN 'ATENDIDO' THEN 2 WHEN 'NO ATENDIDO' THEN 3 ELSE 4 END,v.id ASC"
    c = conn()
    data = c.execute('''SELECT v.*,u.name registered_by,au.name attended_by,su.name status_updated_by
                        FROM visitors v LEFT JOIN users u ON u.id=v.user_id
                        LEFT JOIN users au ON au.id=v.attended_by_user_id
                        LEFT JOIN users su ON su.id=v.status_updated_by_user_id''' + where + order, tuple(params)).fetchall()
    max_conditions = list(conditions); max_params = list(params)
    if status_filter:
        # El sonido debe considerar todos los estados del alcance, no solo el filtro visual.
        max_conditions = [x for x in max_conditions if x != 'v.visit_status=?']
        max_params = max_params[:-1]
    max_where = ' WHERE ' + ' AND '.join(max_conditions) if max_conditions else ''
    row = c.execute('SELECT COALESCE(MAX(v.id),0) ultimo FROM visitors v' + max_where, tuple(max_params)).fetchone()
    c.close()
    return render_template('visitors.html', visitors=data, area_actual=area_actual,
                           institute=institute, campus=campus, ultimo_id=row['ultimo'] or 0,
                           status_filter=status_filter, sonido_area=role in ('sae', 'commercial'))


@app.route('/api/visitas_estado')
@login_required
@role_required('admin', 'sae', 'commercial')
def visitas_estado():
    role = session.get('role')
    conditions, params = [], []
    if role in ('sae', 'commercial'):
        area = 'SAE' if role == 'sae' else 'COMERCIAL'
        institute, campus = session_scope()
        conditions = ['destination_area=?', 'institute=?', 'campus=?']
        params = [area, institute, campus]
    else:
        area = request.args.get('area', '').strip().upper()
        institute = normalize_institute(request.args.get('institute'))
        campus = normalize_campus(request.args.get('campus'))
        if area in AREAS:
            conditions.append('destination_area=?'); params.append(area)
        extra, extra_params = scope_filter('', institute, campus)
        conditions.extend(extra); params.extend(extra_params)
    try:
        since_id = int(request.args.get('since', '0') or 0)
    except ValueError:
        since_id = 0
    where = ' WHERE ' + ' AND '.join(conditions) if conditions else ''
    since_where = where + (' AND ' if where else ' WHERE ') + 'id>?'
    c = conn()
    row = c.execute('''SELECT COUNT(*) total,COALESCE(MAX(id),0) ultimo,
                       SUM(CASE WHEN visit_status='NUEVA' THEN 1 ELSE 0 END) pendientes
                       FROM visitors''' + where, tuple(params)).fetchone()
    nuevos = c.execute('SELECT COUNT(*) n FROM visitors' + since_where, tuple(params + [since_id])).fetchone()['n']
    c.close()
    return jsonify(total=row['total'] or 0, ultimo=row['ultimo'] or 0,
                   pendientes=row['pendientes'] or 0, nuevos=nuevos or 0)


@app.route('/visitas/<int:visit_id>/estado', methods=['POST'])
@login_required
@role_required('admin', 'sae', 'commercial')
def update_visit_status(visit_id):
    estado = request.form.get('estado', 'NUEVA').strip().upper()
    if estado not in ('NUEVA', 'ATENDIDO', 'NO ATENDIDO'):
        estado = 'NUEVA'
    c = conn(); visit = c.execute('SELECT * FROM visitors WHERE id=?', (visit_id,)).fetchone()
    if not visit:
        c.close(); flash('La visita no existe.', 'danger'); return redirect(url_for('visitors'))
    role = session.get('role')
    if role != 'admin':
        expected_area = 'SAE' if role == 'sae' else 'COMERCIAL'
        institute, campus = session_scope()
        if visit['destination_area'] != expected_area or visit['institute'] != institute or visit['campus'] != campus:
            c.close(); flash('No tienes permiso para gestionar visitas de otra sede.', 'danger'); return redirect(url_for('visitors'))
    attended = 1 if estado == 'ATENDIDO' else 0
    attended_by = session['user_id'] if estado == 'ATENDIDO' else None
    status_by = session['user_id'] if estado in ('ATENDIDO', 'NO ATENDIDO') else None
    c.execute('''UPDATE visitors SET visit_status=?,attended=?,attended_at=?,attended_by_user_id=?,
                 status_updated_by_user_id=? WHERE id=?''',
              (estado, attended, now_iso() if attended else None, attended_by, status_by, visit_id))
    c.commit(); c.close()
    flash('Status actualizado y responsable registrado.', 'success')
    redirect_args = {'status': request.form.get('return_status', '')}
    if role == 'admin':
        redirect_args.update(institute=visit['institute'], campus=visit['campus'], area=visit['destination_area'])
    return redirect(url_for('visitors', **redirect_args))


def dashboard_scope():
    institute = normalize_institute(request.args.get('institute'))
    campus = normalize_campus(request.args.get('campus'))
    if institute and campus and not valid_branch(institute, campus):
        campus = ''
    return institute, campus


@app.route('/dashboard')
@login_required
@admin_required
def dashboard():
    institute, campus = dashboard_scope()
    c = conn(); today = today_lima()
    conditions, params = scope_filter('', institute, campus)
    where = ' WHERE ' + ' AND '.join(conditions) if conditions else ''
    today_where = where + (' AND ' if where else ' WHERE ') + 'substr(created_at,1,10)=?'
    today_params = params + [today]

    stats = {
        'active': c.execute('SELECT COUNT(*) n FROM students' + where + (' AND active=1' if where else ' WHERE active=1'), tuple(params)).fetchone()['n'],
        'users': c.execute("SELECT COUNT(*) n FROM users" + where + (" AND role<>'admin' AND active=1" if where else " WHERE role<>'admin' AND active=1"), tuple(params)).fetchone()['n'],
        'visitors_today': c.execute('SELECT COUNT(*) n FROM visitors' + today_where, tuple(today_params)).fetchone()['n'],
        'pending': c.execute("SELECT COUNT(*) n FROM visitors" + where + (" AND visit_status='NUEVA'" if where else " WHERE visit_status='NUEVA'"), tuple(params)).fetchone()['n'],
        'attended': c.execute("SELECT COUNT(*) n FROM visitors" + where + (" AND visit_status='ATENDIDO'" if where else " WHERE visit_status='ATENDIDO'"), tuple(params)).fetchone()['n'],
        'registration_requests': c.execute("SELECT COUNT(*) n FROM registration_requests" + where + (" AND status='PENDIENTE'" if where else " WHERE status='PENDIENTE'"), tuple(params)).fetchone()['n'],
        'branches': 1 if institute and campus else (len(BRANCH_CAMPUSES.get(institute, ())) if institute else len(BRANCHES)),
    }

    user_conditions, user_params = scope_filter('u', institute, campus)
    user_where = ' WHERE ' + ' AND '.join(user_conditions) if user_conditions else ''
    user_activity = c.execute('''SELECT u.name,u.role,u.account_code,u.institute,u.campus,u.active,
        COALESCE(a.alumnos,0) alumnos,COALESCE(t.docentes,0) docentes,
        COALESCE(v.visitas,0) visitas,COALESCE(h.atenciones,0) atenciones,
        COALESCE(a.alumnos,0)+COALESCE(t.docentes,0)+COALESCE(v.visitas,0)+COALESCE(h.atenciones,0) total
        FROM users u
        LEFT JOIN (SELECT user_id,COUNT(*) alumnos FROM access_logs GROUP BY user_id) a ON a.user_id=u.id
        LEFT JOIN (SELECT user_id,COUNT(*) docentes FROM teacher_logs GROUP BY user_id) t ON t.user_id=u.id
        LEFT JOIN (SELECT user_id,COUNT(*) visitas FROM visitors GROUP BY user_id) v ON v.user_id=u.id
        LEFT JOIN (SELECT attended_by_user_id user_id,COUNT(*) atenciones FROM visitors WHERE attended_by_user_id IS NOT NULL GROUP BY attended_by_user_id) h ON h.user_id=u.id''' +
        user_where + " AND u.role<>'admin'" if user_where else '''SELECT u.name,u.role,u.account_code,u.institute,u.campus,u.active,
        COALESCE(a.alumnos,0) alumnos,COALESCE(t.docentes,0) docentes,
        COALESCE(v.visitas,0) visitas,COALESCE(h.atenciones,0) atenciones,
        COALESCE(a.alumnos,0)+COALESCE(t.docentes,0)+COALESCE(v.visitas,0)+COALESCE(h.atenciones,0) total
        FROM users u
        LEFT JOIN (SELECT user_id,COUNT(*) alumnos FROM access_logs GROUP BY user_id) a ON a.user_id=u.id
        LEFT JOIN (SELECT user_id,COUNT(*) docentes FROM teacher_logs GROUP BY user_id) t ON t.user_id=u.id
        LEFT JOIN (SELECT user_id,COUNT(*) visitas FROM visitors GROUP BY user_id) v ON v.user_id=u.id
        LEFT JOIN (SELECT attended_by_user_id user_id,COUNT(*) atenciones FROM visitors WHERE attended_by_user_id IS NOT NULL GROUP BY attended_by_user_id) h ON h.user_id=u.id
        WHERE u.role<>'admin' ''', tuple(user_params)).fetchall()
    # Ordenar en Python evita duplicar una consulta compleja para SQLite/PostgreSQL.
    user_activity = sorted(user_activity, key=lambda r: (-(r['total'] or 0), r['name']))[:30]

    visitor_conditions, visitor_params = scope_filter('v', institute, campus)
    visitor_where = ' WHERE ' + ' AND '.join(visitor_conditions) if visitor_conditions else ''
    recent_visitors = c.execute('''SELECT v.*,ru.name registered_by,au.name attended_by FROM visitors v
                                   LEFT JOIN users ru ON ru.id=v.user_id LEFT JOIN users au ON au.id=v.attended_by_user_id''' +
                                visitor_where + ' ORDER BY v.id DESC LIMIT 20', tuple(visitor_params)).fetchall()

    branch_cards = []
    for inst, sede in BRANCHES:
        if institute and inst != institute:
            continue
        if campus and sede != campus:
            continue
        branch_cards.append({
            'institute': inst, 'campus': sede,
            'students': c.execute('SELECT COUNT(*) n FROM students WHERE active=1 AND institute=? AND campus=?', (inst, sede)).fetchone()['n'],
            'users': c.execute("SELECT COUNT(*) n FROM users WHERE active=1 AND role<>'admin' AND institute=? AND campus=?", (inst, sede)).fetchone()['n'],
            'visits': c.execute("SELECT COUNT(*) n FROM visitors WHERE institute=? AND campus=? AND substr(created_at,1,10)=?", (inst, sede, today)).fetchone()['n'],
            'pending': c.execute("SELECT COUNT(*) n FROM visitors WHERE institute=? AND campus=? AND visit_status='NUEVA'", (inst, sede)).fetchone()['n'],
            'registration_requests': c.execute("SELECT COUNT(*) n FROM registration_requests WHERE institute=? AND campus=? AND status='PENDIENTE'", (inst, sede)).fetchone()['n'],
        })

    charts = {
        'SAE': build_activity_charts(c, 'SAE', institute, campus),
        'COMERCIAL': build_activity_charts(c, 'COMERCIAL', institute, campus),
    }
    request_conditions, request_params = scope_filter('r', institute, campus)
    request_where = ' WHERE ' + ' AND '.join(request_conditions) if request_conditions else ''
    recent_requests = c.execute('''SELECT r.*,u.name requester_name FROM registration_requests r
                                   LEFT JOIN users u ON u.id=r.requester_user_id''' + request_where +
                                ' ORDER BY r.id DESC LIMIT 12', tuple(request_params)).fetchall()
    audits = c.execute('''SELECT a.*,u.name admin_name FROM audit_logs a LEFT JOIN users u ON u.id=a.admin_user_id
                          ORDER BY a.id DESC LIMIT 8''').fetchall()
    c.close()
    return render_template('dashboard.html', stats=stats, selected_institute=institute, selected_campus=campus,
                           branch_cards=branch_cards, user_activity=user_activity, recent_visitors=recent_visitors,
                           admin_activity_charts=charts, recent_requests=recent_requests, audits=audits, lima_now=now_dt().strftime('%d/%m/%Y %H:%M'))


@app.route('/panel')
@login_required
@admin_required
def panel():
    return redirect(url_for('dashboard'))


@app.route('/alumnos')
@login_required
@admin_required
def students():
    q = request.args.get('q', '').strip()
    institute = normalize_institute(request.args.get('institute'))
    campus = normalize_campus(request.args.get('campus'))
    conditions, params = [], []
    if q:
        conditions.append('(name LIKE ? OR dni LIKE ? OR code LIKE ?)')
        params.extend([f'%{q}%', f'%{q}%', f'%{q}%'])
    extra, extra_params = scope_filter('', institute, campus)
    conditions.extend(extra); params.extend(extra_params)
    where = ' WHERE ' + ' AND '.join(conditions) if conditions else ''
    c = conn()
    data = c.execute('SELECT * FROM students' + where + ' ORDER BY active DESC,institute,campus,name', tuple(params)).fetchall()
    summary = []
    for inst, sede in BRANCHES:
        summary.append({'institute': inst, 'campus': sede,
                        'total': c.execute('SELECT COUNT(*) n FROM students WHERE active=1 AND institute=? AND campus=?', (inst, sede)).fetchone()['n']})
    c.close()
    return render_template('students.html', students=data, q=q, institute=institute, campus=campus, summary=summary)


@app.route('/docentes')
@login_required
@admin_required
def teachers():
    q = request.args.get('q', '').strip()
    institute = normalize_institute(request.args.get('institute'))
    campus = normalize_campus(request.args.get('campus'))
    conditions, params = [], []
    if q:
        conditions.append('(name LIKE ? OR dni LIKE ? OR area LIKE ? OR email LIKE ?)')
        params.extend([f'%{q}%', f'%{q}%', f'%{q}%', f'%{q}%'])
    extra, extra_params = scope_filter('', institute, campus)
    conditions.extend(extra); params.extend(extra_params)
    where = ' WHERE ' + ' AND '.join(conditions) if conditions else ''
    c = conn(); data = c.execute('SELECT * FROM teachers' + where + ' ORDER BY active DESC,institute,campus,name', tuple(params)).fetchall(); c.close()
    return render_template('teachers.html', teachers=data, q=q, institute=institute, campus=campus)


@app.route('/registros')
@login_required
@role_required('admin', 'security')
def logs():
    q = request.args.get('q', '').strip()
    if session.get('role') == 'admin':
        institute = normalize_institute(request.args.get('institute'))
        campus = normalize_campus(request.args.get('campus'))
    else:
        institute, campus = session_scope()
    conditions, params = [], []
    extra, extra_params = scope_filter('l', institute, campus)
    # En access_logs los nombres de alcance son registered_*
    if institute:
        conditions.append('l.registered_institute=?'); params.append(institute)
    if campus:
        conditions.append('l.registered_campus=?'); params.append(campus)
    if q:
        conditions.append('(l.student_name LIKE ? OR l.student_dni LIKE ? OR l.student_code LIKE ? OR u.name LIKE ?)')
        params.extend([f'%{q}%', f'%{q}%', f'%{q}%', f'%{q}%'])
    where = ' WHERE ' + ' AND '.join(conditions) if conditions else ''
    c = conn()
    data = c.execute('''SELECT l.*,u.name user FROM access_logs l LEFT JOIN users u ON u.id=l.user_id''' +
                     where + ' ORDER BY l.id DESC LIMIT 500', tuple(params)).fetchall()
    c.close()
    return render_template('logs.html', logs=data, q=q, institute=institute, campus=campus)


@app.route('/importar/alumnos', methods=['POST'])
@login_required
@admin_required
def import_excel():
    file = request.files.get('file')
    institute = normalize_institute(request.form.get('institute'))
    campus = normalize_campus(request.form.get('campus'))
    if not file or not file.filename.lower().endswith(('.xlsx', '.xlsm')):
        flash('Sube un Excel válido .xlsx.', 'danger'); return redirect(url_for('dashboard'))
    if not valid_branch(institute, campus):
        flash('Selecciona una combinación válida de instituto y sede antes de importar.', 'danger'); return redirect(url_for('dashboard'))
    path = os.path.join(UPLOADS, secure_filename(f'{institute}_{CAMPUS_CODES[campus]}_{file.filename}'))
    file.save(path)
    try:
        counts = import_students_from_excel(path, institute, campus)
        total = sum(counts.values())
        audit('IMPORTAR_ALUMNOS', 'students', None, f'{branch_label(institute, campus)} · {total} alumnos')
        flash(f'Base actualizada: {total} alumnos activos en {branch_label(institute, campus)}.', 'success')
    except Exception as exc:
        flash(str(exc), 'danger')
    return redirect(url_for('dashboard', institute=institute, campus=campus))


@app.route('/importar/docentes', methods=['POST'])
@login_required
@admin_required
def import_teachers():
    file = request.files.get('file')
    institute = normalize_institute(request.form.get('institute'))
    campus = normalize_campus(request.form.get('campus'))
    if not file or not file.filename.lower().endswith(('.xlsx', '.xlsm')):
        flash('Sube un Excel válido .xlsx.', 'danger'); return redirect(url_for('teachers'))
    if not valid_branch(institute, campus):
        flash('Selecciona una combinación válida de instituto y sede.', 'danger'); return redirect(url_for('teachers'))
    path = os.path.join(UPLOADS, secure_filename(f'docentes_{institute}_{CAMPUS_CODES[campus]}_{file.filename}'))
    file.save(path)
    try:
        count = import_teachers_from_excel(path, institute, campus)
        audit('IMPORTAR_DOCENTES', 'teachers', None, f'{branch_label(institute, campus)} · {count} docentes')
        flash(f'Base de docentes actualizada: {count} activos en {branch_label(institute, campus)}.', 'success')
    except Exception as exc:
        flash(str(exc), 'danger')
    return redirect(url_for('teachers', institute=institute, campus=campus))


def style_sheet(ws):
    orange = PatternFill('solid', fgColor='F26B2F')
    dark = PatternFill('solid', fgColor='172033')
    thin = Side(style='thin', color='D9E0EA')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = orange
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(bottom=thin)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical='top')
    for col in ws.columns:
        max_len = 12
        letter = col[0].column_letter
        for cell in col:
            max_len = max(max_len, len(str(cell.value or '')) + 2)
        ws.column_dimensions[letter].width = min(max_len, 38)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions


def build_report(kind='visitas', area='', institute='', campus='', days=30):
    days = 30 if days not in (7, 30, 90, 365) else days
    start = (now_dt() - timedelta(days=days)).isoformat(timespec='seconds')
    end = now_iso()
    c = conn(); wb = Workbook(); ws = wb.active
    summary = wb.create_sheet('Resumen')
    summary.append(['PDR · Plataforma de Registro', 'Reporte profesional'])
    summary.append(['Periodo', f'Últimos {days} días'])
    summary.append(['Instituto', institute or 'TODOS'])
    summary.append(['Sede', campus or 'TODAS'])
    summary.append(['Área', area or 'TODAS'])
    summary.append(['Generado', now_dt().strftime('%d/%m/%Y %H:%M')])
    summary.column_dimensions['A'].width = 24; summary.column_dimensions['B'].width = 42
    for cell in summary[1]:
        cell.font = Font(bold=True, color='FFFFFF'); cell.fill = PatternFill('solid', fgColor='172033')

    if kind == 'visitas':
        ws.title = 'Visitas'
        ws.append(['Fecha/Hora', 'Instituto', 'Sede', 'Registró', 'Nombre visitante', 'DNI', 'Teléfono',
                   'Área destino', 'Motivo', 'Status', 'Atendió', 'Actualizó status', 'Atendido en'])
        conditions = ['v.created_at BETWEEN ? AND ?']; params = [start, end]
        if area in AREAS:
            conditions.append('v.destination_area=?'); params.append(area)
        extra, extra_params = scope_filter('v', institute, campus)
        conditions.extend(extra); params.extend(extra_params)
        rows = c.execute('''SELECT v.*,ru.name registered_by,au.name attended_by,su.name status_updated_by
                            FROM visitors v LEFT JOIN users ru ON ru.id=v.user_id
                            LEFT JOIN users au ON au.id=v.attended_by_user_id
                            LEFT JOIN users su ON su.id=v.status_updated_by_user_id
                            WHERE ''' + ' AND '.join(conditions) + ' ORDER BY v.created_at DESC', tuple(params)).fetchall()
        for row in rows:
            ws.append([row['created_at'], row['institute'], row['campus'], row['registered_by'] or '', row['full_name'],
                       row['dni'], row['phone'] or '', row['destination_area'], row['reason'] or row['visit_type'] or '',
                       row['visit_status'], row['attended_by'] or '', row['status_updated_by'] or '', row['attended_at'] or ''])
        filename = f'PDR_visitas_{institute or "TODOS"}_{CAMPUS_CODES.get(campus,"TODAS")}_{now_dt().strftime("%Y%m%d_%H%M")}.xlsx'
    else:
        ws.title = 'Alumnos'
        ws.append(['Fecha/Hora', 'Instituto registro', 'Sede registro', 'Registró', 'Consulta', 'Resultado',
                   'Alumno', 'Instituto alumno', 'Sede alumno', 'DNI', 'Código', 'Nota'])
        conditions = ['l.created_at BETWEEN ? AND ?']; params = [start, end]
        if institute:
            conditions.append('l.registered_institute=?'); params.append(institute)
        if campus:
            conditions.append('l.registered_campus=?'); params.append(campus)
        rows = c.execute('''SELECT l.*,u.name registered_by FROM access_logs l LEFT JOIN users u ON u.id=l.user_id
                            WHERE ''' + ' AND '.join(conditions) + ' ORDER BY l.created_at DESC', tuple(params)).fetchall()
        for row in rows:
            ws.append([row['created_at'], row['registered_institute'], row['registered_campus'], row['registered_by'] or '',
                       row['query'], row['result'], row['student_name'] or '', row['student_institute'] or '',
                       row['student_campus'] or '', row['student_dni'] or '', row['student_code'] or '', row['note'] or ''])
        filename = f'PDR_registros_{institute or "TODOS"}_{CAMPUS_CODES.get(campus,"TODAS")}_{now_dt().strftime("%Y%m%d_%H%M")}.xlsx'
    c.close(); style_sheet(ws)
    path = os.path.join(UPLOADS, filename); wb.save(path)
    return path


@app.route('/exportar/solicitudes')
@login_required
@role_required('admin', 'security')
def export_registration_requests():
    if session.get('role') == 'admin':
        institute = normalize_institute(request.args.get('institute'))
        campus = normalize_campus(request.args.get('campus'))
        if institute and campus and not valid_branch(institute, campus):
            campus = ''
    else:
        institute, campus = session_scope()
    conditions, params = [], []
    if institute:
        conditions.append('r.institute=?'); params.append(institute)
    if campus:
        conditions.append('r.campus=?'); params.append(campus)
    where = ' WHERE ' + ' AND '.join(conditions) if conditions else ''
    c = conn()
    rows = c.execute('''SELECT r.*,u.name requester_name,rv.name reviewer_name
                        FROM registration_requests r
                        LEFT JOIN users u ON u.id=r.requester_user_id
                        LEFT JOIN users rv ON rv.id=r.reviewed_by_user_id''' + where +
                     ' ORDER BY r.created_at DESC', tuple(params)).fetchall()
    c.close()
    wb = Workbook(); ws = wb.active; ws.title = 'Solicitudes'
    ws.append(['Fecha', 'Instituto', 'Sede', 'Tipo', 'Nombre', 'DNI', 'Código', 'Teléfono',
               'Detalle', 'Solicitó', 'Estado', 'Revisó', 'Fecha revisión'])
    for row in rows:
        ws.append([row['created_at'], row['institute'], row['campus'], row['request_type'], row['full_name'],
                   row['dni'], row['code'] or '', row['phone'] or '', row['detail'] or '',
                   row['requester_name'] or '', row['status'], row['reviewer_name'] or '', row['reviewed_at'] or ''])
    style_sheet(ws)
    filename = f'PDR_solicitudes_{institute or "TODOS"}_{CAMPUS_CODES.get(campus,"TODAS")}_{now_dt().strftime("%Y%m%d_%H%M")}.xlsx'
    path = os.path.join(UPLOADS, filename); wb.save(path)
    return send_file(path, as_attachment=True)


@app.route('/exportar')
@login_required
@role_required('admin', 'security')
def export_excel():
    if session.get('role') == 'admin':
        institute = normalize_institute(request.args.get('institute'))
        campus = normalize_campus(request.args.get('campus'))
    else:
        institute, campus = session_scope()
    try:
        days = int(request.args.get('days', '30'))
    except ValueError:
        days = 30
    return send_file(build_report('registros', institute=institute, campus=campus, days=days), as_attachment=True)


@app.route('/exportar/visitas')
@login_required
@role_required('admin', 'sae', 'commercial')
def export_visits_excel():
    role = session.get('role')
    if role == 'admin':
        institute = normalize_institute(request.args.get('institute'))
        campus = normalize_campus(request.args.get('campus'))
        area = request.args.get('area', '').strip().upper()
        if area not in AREAS:
            area = ''
    else:
        institute, campus = session_scope()
        area = 'SAE' if role == 'sae' else 'COMERCIAL'
    try:
        days = int(request.args.get('days', '30'))
    except ValueError:
        days = 30
    return send_file(build_report('visitas', area=area, institute=institute, campus=campus, days=days), as_attachment=True)


init_db()

if __name__ == '__main__':
    app.run(debug=True)
